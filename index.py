"""
api/index.py  —  Coretax DLP XML  →  ImporPajakKeluaran XLSX / CSV
Vercel Serverless Function (Flask).
Untuk local dev: python index.py
"""

from flask import Flask, request, send_file, jsonify, Response
import io, csv, math, random
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import datetime
from xml.etree import ElementTree as ET
from openpyxl import load_workbook

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

COUNTRY_FIX = {"IND": "IDN"}

OPT_MAP = {
    "A": "GOODS", "B": "GOODS",
    "J": "SERVICES", "S": "SERVICES",
    "GOODS": "GOODS", "SERVICES": "SERVICES",
}

DOMESTIC_DOC = {"tin", "npwp", "nik", ""}


def _text(el, tag):
    child = el.find(tag)
    if child is None:
        return None
    t = (child.text or "").strip()
    return t if t else None


def _num(el, tag):
    """Parse XML text to Decimal (exact). Returns None if missing/invalid."""
    raw = _text(el, tag)
    if raw is None:
        return None
    cleaned = raw.strip().replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _to_decimal(v):
    """Safely coerce any value (float, int, str, Decimal) to Decimal. Returns None on failure."""
    if v is None:
        return None
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return None


def _round_no_decimal(v):
    d = _to_decimal(v)
    if d is None:
        return None
    floor_d = d.to_integral_value(rounding="ROUND_FLOOR")
    frac = d - floor_d
    if frac >= Decimal("0.50"):
        return int(floor_d) + 1
    else:
        return int(floor_d)


def to_int_or_float(v):
    if v is None:
        return None
    return _round_no_decimal(v)


def fmt_str(v):
    if v is None:
        return None
    return str(_round_no_decimal(v))


def format_date(raw):
    try:
        return datetime.strptime(raw.strip(), "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return raw or None


def get_month(raw):
    try:
        return datetime.strptime(raw.strip(), "%Y-%m-%d").month
    except Exception:
        return None


def get_year(raw):
    try:
        return datetime.strptime(raw.strip(), "%Y-%m-%d").strftime("%Y")
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# XML Parsing
# ─────────────────────────────────────────────────────────────────────────────

def parse_xml_bytes(xml_bytes):
    root       = ET.fromstring(xml_bytes)
    seller_tin = _text(root, "TIN")
    invoices   = []

    for inv in root.iter("TaxInvoice"):
        raw_date    = _text(inv, "TaxInvoiceDate") or ""
        opt         = (_text(inv, "TaxInvoiceOpt") or "Normal").strip()
        buyer_doc   = (_text(inv, "BuyerDocument") or "").strip().lower()
        buyer_docno = _text(inv, "BuyerDocumentNumber")

        fg_pengganti = 1 if opt.lower() == "pengganti" else 0

        is_domestic = buyer_doc in DOMESTIC_DOC
        raw_country = _text(inv, "BuyerCountry") or ""
        kode_negara = None if is_domestic else (
            COUNTRY_FIX.get(raw_country.upper(), raw_country) or None
        )

        invoice = {
            "seller_tin":           seller_tin,
            "kd_jenis_transaksi":   _text(inv, "TrxCode"),
            "fg_pengganti":         fg_pengganti,
            "masa_pajak":           get_month(raw_date),
            "tahun_pajak":          get_year(raw_date),
            "tanggal_faktur":       format_date(raw_date),
            "npwp":                 _text(inv, "BuyerTin"),
            "nama":                 _text(inv, "BuyerName"),
            "alamat_lengkap":       _text(inv, "BuyerAdress"),
            "id_keterangan_tambah": _text(inv, "AddInfo"),
            "referensi":            _text(inv, "RefDesc"),
            "kode_dok_pendukung":   _text(inv, "CustomDoc"),
            "passport":             buyer_docno if "paspor" in buyer_doc else None,
            "id_lain":              buyer_docno if (
                buyer_doc not in DOMESTIC_DOC and "paspor" not in buyer_doc
            ) else None,
            "kode_negara":          kode_negara,
            "id_tku_penjual":       _text(inv, "SellerIDTKU"),
            "nomor_faktur_diganti": _text(inv, "ReplacedTaxInvoiceNo") if fg_pengganti else None,
            "email":                _text(inv, "BuyerEmail"),
            "keterangan1":          _text(inv, "FacilityStamp"),
        }

        goods = []
        for gs in inv.iter("GoodService"):
            price    = _num(gs, "Price")
            qty      = _num(gs, "Qty")
            tax_base = _num(gs, "TaxBase")
            discount = _num(gs, "TotalDiscount")
            harga_total = (price * qty) if (price is not None and qty is not None) else None
            dpp_val = ((tax_base or 0) - (discount or 0)) if tax_base is not None else None
            goods.append({
                "kode_objek":     _text(gs, "Code"),
                "nama":           _text(gs, "Name"),
                "harga_satuan":   to_int_or_float(price),
                "jumlah_barang":  to_int_or_float(qty),
                "harga_total":    to_int_or_float(harga_total),
                "diskon":         to_int_or_float(discount),
                "dpp":            to_int_or_float(dpp_val),
                "ppn":            to_int_or_float(_num(gs, "VAT")),
                "tarif_ppnbm":    fmt_str(_num(gs, "STLGRate")),
                "ppnbm":          fmt_str(_num(gs, "STLG")),
                "brgjasa":        OPT_MAP.get((_text(gs, "Opt") or "").strip().upper(), "GOODS"),
                "satuanbrgjasa":  _text(gs, "Unit"),
                "dpp_nilai_lain": to_int_or_float(_num(gs, "OtherTaxBase")),
            })

        def total_str(key):
            vals = [g[key] for g in goods if g[key] is not None]
            if not vals:
                return None
            return str(sum(int(v) for v in vals))

        invoice["jumlah_dpp"]   = total_str("dpp")
        invoice["jumlah_ppn"]   = total_str("ppn")
        invoice["jumlah_ppnbm"] = total_str("ppnbm")
        invoices.append((invoice, goods))

    return invoices


# ─────────────────────────────────────────────────────────────────────────────
# NOMOR_FAKTUR Generator — random 5-digit, unique per session
# ─────────────────────────────────────────────────────────────────────────────

def gen_nomor_random(used_set):
    """Generate a unique random 5-digit NO_FAKTUR (10000–99999)."""
    while True:
        num = random.randint(10000, 99999)
        if num not in used_set:
            used_set.add(num)
            return str(num)


# ─────────────────────────────────────────────────────────────────────────────
# Row Builder
# ─────────────────────────────────────────────────────────────────────────────

def build_data_rows(invoices):
    rows = []
    used = set()
    for inv, goods in invoices:
        nomor = gen_nomor_random(used)
        fk = [
            "FK",
            inv["kd_jenis_transaksi"],
            inv["fg_pengganti"],
            nomor,
            inv["masa_pajak"],
            inv["tahun_pajak"],
            inv["tanggal_faktur"],
            inv["npwp"],
            inv["nama"],
            inv["alamat_lengkap"],
            inv["jumlah_dpp"],
            inv["jumlah_ppn"],
            inv["jumlah_ppnbm"],
            inv["id_keterangan_tambah"],
            None, None, None, None, None,
            inv["referensi"],
            inv["kode_dok_pendukung"],
            None,
            inv["passport"],
            inv["id_lain"],
            inv["kode_negara"],
            inv["id_tku_penjual"],
            inv["nomor_faktur_diganti"],
            inv["email"],
            inv["keterangan1"],
            None, None, None, None,
        ]
        rows.append(fk)

        for gs in goods:
            of = [
                "OF",
                gs["kode_objek"],
                gs["nama"],
                gs["harga_satuan"],
                gs["jumlah_barang"],
                gs["harga_total"],
                gs["diskon"],
                gs["dpp"],
                gs["ppn"],
                gs["tarif_ppnbm"],
                gs["ppnbm"],
                gs["brgjasa"],
                gs["satuanbrgjasa"],
                gs["dpp_nilai_lain"],
            ] + [None] * 19
            rows.append(of)

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Writers
# ─────────────────────────────────────────────────────────────────────────────

def write_xlsx(invoices, template_bytes):
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb["DATA"]
    if ws.max_row and ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row - 3)
    for row in build_data_rows(invoices):
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


_H_FK = [
    "FK","KD_JENIS_TRANSAKSI","FG_PENGGANTI","NOMOR_FAKTUR","MASA_PAJAK",
    "TAHUN_PAJAK","TANGGAL_FAKTUR","NPWP","NAMA","ALAMAT_LENGKAP",
    "JUMLAH_DPP","JUMLAH_PPN","JUMLAH_PPNBM","ID_KETERANGAN_TAMBAH",
    "FG_UANG_MUKA","UANG_MUKA_DPP","UANG_MUKA_PPN","UANG_MUKA_PPNBM",
    "UANG_MUKA_DPP_LAIN","REFERENSI","KODE_DOKUMEN_PENDUKUNG",
    "NOMOR_FAKTUR_UANG_MUKA","PASSPORT","ID_LAIN","KODE_NEGARA",
    "ID_TKU_PENJUAL","NOMOR_FAKTUR_DIGANTI","EMAIL",
    "KETERANGAN1","KETERANGAN2","KETERANGAN3","KETERANGAN4","KETERANGAN5",
]
_H_LT = (["LT","NPWP","NAMA","JALAN","BLOK","NOMOR","RT","RW",
           "KECAMATAN","KELURAHAN","KABUPATEN","PROPINSI","KODE_POS","NOMOR_TELEPON"]
          + [""] * 19)
_H_OF = (["OF","KODE_OBJEK","NAMA","HARGA_SATUAN","JUMLAH_BARANG","HARGA_TOTAL",
           "DISKON","DPP","PPN","TARIF_PPNBM","PPNBM","BRGJASA","SATUANBRGJASA",
           "DPP_NILAI_LAIN"] + [""] * 19)


def write_csv(invoices):
    buf = io.StringIO()
    w   = csv.writer(buf, lineterminator="\r\n")
    w.writerow(_H_FK)
    w.writerow(_H_LT)
    w.writerow(_H_OF)
    for row in build_data_rows(invoices):
        w.writerow(["" if v is None else v for v in row])
    return io.BytesIO(buf.getvalue().encode("utf-8-sig"))


# ─────────────────────────────────────────────────────────────────────────────
# HTML — friendly, warm, bright design
# ─────────────────────────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Coretax → PajakExpress Converter</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
:root {
  --bg: #f0f4ff;
  --card: #ffffff;
  --border: #dde4f5;
  --accent: #4f6ef7;
  --accent2: #7c3aed;
  --green: #10b981;
  --yellow: #f59e0b;
  --red: #ef4444;
  --text: #1e293b;
  --sub: #64748b;
  --muted: #94a3b8;
  --shadow: 0 4px 24px rgba(79,110,247,.10);
  --shadow-hover: 0 8px 32px rgba(79,110,247,.18);
}
* { box-sizing:border-box; margin:0; padding:0; }
body {
  background: var(--bg);
  background-image:
    radial-gradient(circle at 15% 15%, rgba(124,58,237,.10) 0%, transparent 50%),
    radial-gradient(circle at 85% 80%, rgba(79,110,247,.10) 0%, transparent 50%);
  color: var(--text);
  font-family: 'Nunito', sans-serif;
  min-height: 100vh;
  display: flex;
  flex-direction: column;
  align-items: center;
  padding: 40px 16px 80px;
}

.wrap { width: 100%; max-width: 620px; }

/* ── Header ── */
header { text-align: center; margin-bottom: 36px; }
.badge {
  display: inline-block;
  background: linear-gradient(135deg, var(--accent), var(--accent2));
  color: #fff;
  font-size: 10px;
  font-weight: 800;
  letter-spacing: 3px;
  padding: 4px 14px;
  border-radius: 99px;
  margin-bottom: 14px;
  text-transform: uppercase;
}
h1 {
  font-size: 32px;
  font-weight: 800;
  color: var(--text);
  line-height: 1.1;
}
h1 .arrow {
  display: inline-block;
  background: linear-gradient(135deg, var(--accent), var(--accent2));
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  background-clip: text;
  padding: 0 6px;
}
.sub-h {
  font-size: 14px;
  color: var(--sub);
  margin-top: 8px;
  font-weight: 600;
}

/* ── Cards ── */
.card {
  background: var(--card);
  border: 1.5px solid var(--border);
  border-radius: 16px;
  padding: 24px;
  margin-bottom: 16px;
  box-shadow: var(--shadow);
  transition: box-shadow .2s;
}
.card:hover { box-shadow: var(--shadow-hover); }

.section-label {
  display: flex;
  align-items: center;
  gap: 8px;
  font-size: 11px;
  font-weight: 800;
  letter-spacing: 2px;
  text-transform: uppercase;
  color: var(--accent);
  margin-bottom: 16px;
}
.section-label .num {
  background: linear-gradient(135deg, var(--accent), var(--accent2));
  color: #fff;
  width: 22px; height: 22px;
  border-radius: 6px;
  display: flex; align-items: center; justify-content: center;
  font-size: 11px; font-weight: 800;
}

/* ── Fields ── */
.field + .field { margin-top: 14px; }
label {
  display: block;
  font-size: 13px;
  font-weight: 700;
  color: var(--text);
  margin-bottom: 6px;
}
label .opt { font-weight: 400; color: var(--muted); font-size: 12px; }

.file-zone {
  border: 2px dashed var(--border);
  border-radius: 12px;
  padding: 20px;
  text-align: center;
  cursor: pointer;
  transition: all .2s;
  position: relative;
  background: #fafbff;
}
.file-zone:hover, .file-zone.has-file {
  border-color: var(--accent);
  background: rgba(79,110,247,.04);
}
.file-zone input[type=file] {
  position: absolute; inset: 0; opacity: 0; cursor: pointer; width: 100%;
}
.file-zone .icon { font-size: 28px; margin-bottom: 6px; }
.file-zone .fz-label { font-size: 13px; font-weight: 700; color: var(--sub); }
.file-zone .fz-sub { font-size: 11px; color: var(--muted); margin-top: 3px; }
.file-zone.has-file .fz-label { color: var(--accent); }

/* ── Info box ── */
.info-box {
  background: linear-gradient(135deg, rgba(79,110,247,.07), rgba(124,58,237,.07));
  border: 1.5px solid rgba(79,110,247,.2);
  border-radius: 10px;
  padding: 14px 16px;
  font-size: 13px;
  color: var(--sub);
  line-height: 1.6;
  font-weight: 600;
}
.info-box .tag {
  display: inline-block;
  background: var(--accent);
  color: #fff;
  font-size: 10px;
  font-weight: 800;
  padding: 2px 8px;
  border-radius: 5px;
  font-family: 'JetBrains Mono', monospace;
  margin-left: 4px;
}

/* ── Format toggle ── */
.fmt-toggle { display: flex; gap: 10px; }
.fmt-btn {
  flex: 1;
  border: 2px solid var(--border);
  border-radius: 12px;
  padding: 14px 10px;
  cursor: pointer;
  transition: all .2s;
  text-align: center;
  background: #fafbff;
  font-family: 'Nunito', sans-serif;
  font-size: 14px;
  font-weight: 700;
  color: var(--sub);
  user-select: none;
}
.fmt-btn .fmt-icon { font-size: 22px; display: block; margin-bottom: 4px; }
.fmt-btn:hover { border-color: var(--accent); color: var(--accent); background: rgba(79,110,247,.04); }
.fmt-btn.active {
  border-color: var(--accent);
  color: var(--accent);
  background: rgba(79,110,247,.08);
  box-shadow: 0 0 0 3px rgba(79,110,247,.15);
}

/* ── Stats ── */
.stats { display: flex; gap: 10px; margin-bottom: 14px; }
.stat {
  flex: 1;
  border-radius: 12px;
  padding: 16px;
  text-align: center;
  background: linear-gradient(135deg, #f0f4ff, #e8edff);
  border: 1.5px solid var(--border);
}
.stat-num {
  font-size: 30px;
  font-weight: 800;
  color: var(--accent);
  font-family: 'JetBrains Mono', monospace;
  line-height: 1;
}
.stat-num.g { color: var(--green); }
.stat-lbl { font-size: 11px; font-weight: 700; color: var(--sub); margin-top: 4px; text-transform: uppercase; letter-spacing: 1px; }

/* ── Log ── */
.log {
  background: #0f172a;
  border-radius: 10px;
  font-family: 'JetBrains Mono', monospace;
  font-size: 11px;
  color: #94a3b8;
  padding: 14px;
  height: 140px;
  overflow-y: auto;
  white-space: pre-wrap;
  word-break: break-all;
  line-height: 1.6;
}
.log .ok   { color: #34d399; }
.log .err  { color: #f87171; }
.log .warn { color: #fbbf24; }

/* ── Button ── */
.btn {
  width: 100%;
  background: linear-gradient(135deg, var(--accent), var(--accent2));
  color: #fff;
  border: none;
  border-radius: 14px;
  font-family: 'Nunito', sans-serif;
  font-size: 16px;
  font-weight: 800;
  padding: 18px;
  cursor: pointer;
  transition: all .2s;
  box-shadow: 0 6px 20px rgba(79,110,247,.35);
  letter-spacing: .5px;
  margin-top: 4px;
}
.btn:hover { transform: translateY(-2px); box-shadow: 0 10px 28px rgba(79,110,247,.45); }
.btn:active { transform: translateY(0); }
.btn:disabled { opacity: .5; cursor: not-allowed; transform: none; box-shadow: none; }

/* ── Banner ── */
.banner {
  display: none;
  background: linear-gradient(135deg, rgba(16,185,129,.1), rgba(16,185,129,.05));
  border: 1.5px solid rgba(16,185,129,.4);
  border-radius: 12px;
  padding: 14px 18px;
  color: var(--green);
  font-size: 14px;
  font-weight: 700;
  margin-bottom: 16px;
  text-align: center;
}

/* ── Footer ── */
footer {
  margin-top: 40px;
  font-size: 12px;
  color: var(--muted);
  text-align: center;
  font-weight: 600;
}
</style>
</head>
<body>
<div class="wrap">

  <header>
    <div class="badge">✦ Pajak Converter ✦</div>
    <h1>XML <span class="arrow">→</span> PajakExpress</h1>
    <p class="sub-h">Konversi Coretax DLP XML ke format ImporPajakKeluaran XLSX / CSV</p>
  </header>

  <div id="banner" class="banner"></div>

  <!-- 01 FILES -->
  <div class="card">
    <div class="section-label"><span class="num">1</span> Upload File</div>

    <div class="field">
      <label>File XML Coretax <span style="color:var(--red)">*</span></label>
      <div class="file-zone" id="xml-zone">
        <input type="file" id="xml-file" accept=".xml">
        <div class="icon">📄</div>
        <div class="fz-label" id="xml-label">Klik atau drag file XML ke sini</div>
        <div class="fz-sub">Format: .xml dari Coretax DLP</div>
      </div>
    </div>

    <div class="field" id="tmpl-wrap">
      <label>Template XLSX <span class="opt">(wajib untuk output XLSX)</span></label>
      <div class="file-zone" id="tmpl-zone">
        <input type="file" id="tmpl-file" accept=".xlsx">
        <div class="icon">📊</div>
        <div class="fz-label" id="tmpl-label">Klik atau drag template XLSX ke sini</div>
        <div class="fz-sub">ImporPajakKeluaran_CSV_.xlsx</div>
      </div>
    </div>
  </div>

  <!-- 02 NO FAKTUR INFO -->
  <div class="card">
    <div class="section-label"><span class="num">2</span> Nomor Faktur</div>
    <div class="info-box">
      🎲 <strong>Auto-generate otomatis</strong> — NO_FAKTUR tidak tersedia di XML Coretax,
      sehingga sistem akan membuat <strong>nomor acak 5 digit unik</strong>
      <span class="tag">10000–99999</span> untuk setiap faktur secara otomatis.
      Tidak perlu pengaturan tambahan!
    </div>
  </div>

  <!-- 03 FORMAT -->
  <div class="card">
    <div class="section-label"><span class="num">3</span> Format Output</div>
    <div class="fmt-toggle">
      <div class="fmt-btn active" id="btn-xlsx">
        <span class="fmt-icon">📊</span>
        XLSX
      </div>
      <div class="fmt-btn" id="btn-csv">
        <span class="fmt-icon">📋</span>
        CSV
      </div>
    </div>
    <p style="font-size:12px; color:var(--sub); margin-top:12px; font-weight:600;" id="fmt-hint">
      XLSX: mengisi template ImporPajakKeluaran (perlu upload template di atas).
    </p>
  </div>

  <!-- 04 PREVIEW -->
  <div class="card">
    <div class="section-label"><span class="num">4</span> Preview</div>
    <div class="stats">
      <div class="stat">
        <div class="stat-num" id="inv-count">—</div>
        <div class="stat-lbl">Faktur (FK)</div>
      </div>
      <div class="stat">
        <div class="stat-num g" id="item-count">—</div>
        <div class="stat-lbl">Item (OF)</div>
      </div>
    </div>
    <div class="log" id="log">Pilih file XML untuk melihat preview…</div>
  </div>

  <button class="btn" id="conv-btn">▶ &nbsp;Convert &amp; Download</button>

  <footer>Coretax DLP XML → ImporPajakKeluaran &nbsp;·&nbsp; XLSX &amp; CSV</footer>
</div>

<script>
(function () {
  var outputFmt = 'xlsx';

  /* ── File zone labels ── */
  function setupFileZone(inputId, zoneId, labelId) {
    var input = document.getElementById(inputId);
    var zone  = document.getElementById(zoneId);
    var lbl   = document.getElementById(labelId);
    input.addEventListener('change', function () {
      if (this.files.length) {
        lbl.textContent = this.files[0].name;
        zone.classList.add('has-file');
      }
    });
  }
  setupFileZone('xml-file',  'xml-zone',  'xml-label');
  setupFileZone('tmpl-file', 'tmpl-zone', 'tmpl-label');

  /* ── Format toggle ── */
  document.getElementById('btn-xlsx').addEventListener('click', function () { setFmt('xlsx'); });
  document.getElementById('btn-csv').addEventListener('click',  function () { setFmt('csv');  });

  function setFmt(f) {
    outputFmt = f;
    document.getElementById('btn-xlsx').classList.toggle('active', f === 'xlsx');
    document.getElementById('btn-csv').classList.toggle('active',  f === 'csv');
    document.getElementById('tmpl-wrap').style.display = (f === 'xlsx') ? '' : 'none';
    document.getElementById('fmt-hint').textContent = (f === 'xlsx')
      ? 'XLSX: mengisi template ImporPajakKeluaran (perlu upload template di atas).'
      : 'CSV: UTF-8 BOM, flat file, langsung bisa diupload ke PajakExpress.';
  }

  /* ── Log helpers ── */
  function addLog(msg, cls) {
    var el   = document.getElementById('log');
    var span = document.createElement('span');
    if (cls) span.className = cls;
    span.textContent = msg + '\n';
    el.appendChild(span);
    el.scrollTop = el.scrollHeight;
  }
  function clearLog() { document.getElementById('log').innerHTML = ''; }

  /* ── XML preview ── */
  document.getElementById('xml-file').addEventListener('change', function () {
    if (!this.files.length) return;
    var fd = new FormData();
    fd.append('xml', this.files[0]);
    clearLog();
    addLog('Membaca XML…');

    fetch('/api/preview', { method: 'POST', body: fd })
      .then(function (res) { return res.json(); })
      .then(function (data) {
        if (data.error) { addLog('ERROR: ' + data.error, 'err'); return; }
        document.getElementById('inv-count').textContent  = data.invoices;
        document.getElementById('item-count').textContent = data.items;
        (data.warnings || []).forEach(function (w) { addLog('⚠ ' + w, 'warn'); });
        data.preview.forEach(function (p) { addLog(p, 'ok'); });
      })
      .catch(function (e) { addLog('Gagal: ' + e, 'err'); });
  });

  /* ── Convert ── */
  document.getElementById('conv-btn').addEventListener('click', function () {
    var xmlFile  = document.getElementById('xml-file').files[0];
    var tmplFile = document.getElementById('tmpl-file').files[0];

    if (!xmlFile) { alert('Pilih file XML Coretax terlebih dahulu.'); return; }
    if (outputFmt === 'xlsx' && !tmplFile) {
      alert('Untuk output XLSX, upload dulu file template ImporPajakKeluaran_CSV_.xlsx.\nAtau ganti format ke CSV.');
      return;
    }

    var btn = document.getElementById('conv-btn');
    btn.disabled    = true;
    btn.textContent = '⏳  Sedang Mengonversi…';
    clearLog();
    addLog('Memproses ' + xmlFile.name + '…');

    var fd = new FormData();
    fd.append('xml',    xmlFile);
    fd.append('format', outputFmt);
    if (tmplFile) fd.append('template', tmplFile);

    fetch('/api/convert', { method: 'POST', body: fd })
      .then(function (res) {
        if (!res.ok) {
          return res.json().then(function (err) {
            throw new Error(err.error || res.statusText);
          });
        }
        return res.blob();
      })
      .then(function (blob) {
        var ext  = (outputFmt === 'csv') ? '.csv' : '.xlsx';
        var name = xmlFile.name.replace(/\.xml$/i, '') + '_converted' + ext;
        var url  = URL.createObjectURL(blob);
        var a    = document.createElement('a');
        a.href = url; a.download = name; a.style.display = 'none';
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        setTimeout(function () { URL.revokeObjectURL(url); }, 500);

        addLog('✓ Selesai — ' + name + ' sudah diunduh!', 'ok');
        var banner = document.getElementById('banner');
        banner.innerHTML = '🎉 Konversi berhasil! File <strong>' + name + '</strong> sudah diunduh.';
        banner.style.display = 'block';
      })
      .catch(function (e) {
        addLog('ERROR: ' + e.message, 'err');
      })
      .finally(function () {
        btn.disabled    = false;
        btn.textContent = '▶  Convert & Download';
      });
  });

})();
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return Response(HTML, mimetype="text/html")


@app.route("/api/preview", methods=["POST"])
def preview():
    try:
        invoices    = parse_xml_bytes(request.files["xml"].read())
        total_items = sum(len(g) for _, g in invoices)
        warnings    = []
        lines       = []

        for i, (inv, goods) in enumerate(invoices, 1):
            if not inv.get("npwp"):
                warnings.append(f"Faktur #{i}: NPWP pembeli kosong")
            if not inv.get("kd_jenis_transaksi"):
                warnings.append(f"Faktur #{i}: TrxCode tidak ditemukan di XML")
            lines.append(
                f"[{i}] {inv['nama'] or '-'}  |  "
                f"{inv['tanggal_faktur'] or '-'}  |  "
                f"TrxCode={inv['kd_jenis_transaksi'] or '-'}  |  "
                f"{len(goods)} item"
            )

        return jsonify(invoices=len(invoices), items=total_items,
                       preview=lines, warnings=warnings)
    except Exception as e:
        return jsonify(error=str(e)), 400


@app.route("/api/convert", methods=["POST"])
def convert():
    try:
        invoices = parse_xml_bytes(request.files["xml"].read())
        fmt      = request.form.get("format", "xlsx").lower()

        if fmt == "csv":
            buf     = write_csv(invoices)
            mime    = "text/csv"
            dl_name = "converted.csv"
        else:
            tmpl = request.files.get("template")
            if not tmpl:
                return jsonify(error="Template XLSX wajib disertakan untuk output XLSX."), 400
            buf     = write_xlsx(invoices, tmpl.read())
            mime    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            dl_name = "converted.xlsx"

        return send_file(buf, as_attachment=True, download_name=dl_name, mimetype=mime)

    except Exception as e:
        return jsonify(error=str(e)), 400


if __name__ == "__main__":
    app.run(debug=True, port=5000)
